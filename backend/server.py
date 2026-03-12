from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from enum import Enum
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "your-secret-key-change-in-production-2024")
ALGORITHM = "HS256"

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Enums
class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    BRANCH_MANAGER = "branch_manager"
    CASHIER = "cashier"
    LOAN_OFFICER = "loan_officer"
    SAVINGS_OFFICER = "savings_officer"
    AGENT = "agent"
    POS_OPERATOR = "pos_operator"
    AUDITOR = "auditor"
    INVESTOR = "investor"

class SavingsFrequency(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"

class LoanStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    DISBURSED = "disbursed"
    ACTIVE = "active"
    COMPLETED = "completed"
    DEFAULTED = "defaulted"

class TransactionType(str, Enum):
    DEPOSIT = "deposit"
    WITHDRAWAL = "withdrawal"
    LOAN_DISBURSEMENT = "loan_disbursement"
    LOAN_REPAYMENT = "loan_repayment"
    INVESTMENT = "investment"
    PAYOUT = "payout"
    SALE = "sale"
    EXPENSE = "expense"

class CardType(str, Enum):
    VISA = "visa"
    MASTERCARD = "mastercard"
    VERVE = "verve"
    OTHER = "other"

class ATMTransactionType(str, Enum):
    WITHDRAWAL = "withdrawal"
    DEPOSIT = "deposit"
    BALANCE_INQUIRY = "balance_inquiry"

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    full_name: str
    phone: str
    role: UserRole
    branch_id: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    phone: str
    role: UserRole
    branch_id: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: User

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_number: str
    full_name: str
    phone: str
    address: str
    photo_url: Optional[str] = None
    agent_id: str
    branch_id: str
    guarantor_name: Optional[str] = None
    guarantor_phone: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    full_name: str
    phone: str
    address: str
    guarantor_name: Optional[str] = None
    guarantor_phone: Optional[str] = None

class SavingsAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_number: str
    customer_id: str
    frequency: SavingsFrequency
    target_amount: float
    current_balance: float = 0.0
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SavingsAccountCreate(BaseModel):
    customer_id: str
    frequency: SavingsFrequency
    target_amount: float

class SavingsTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_id: str
    customer_id: str
    agent_id: str
    amount: float
    collection_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    submitted_to_office: bool = False
    verified_by_cashier: bool = False
    cashier_id: Optional[str] = None
    verification_date: Optional[datetime] = None
    notes: Optional[str] = None

class SavingsTransactionCreate(BaseModel):
    account_id: str
    customer_id: str
    amount: float
    notes: Optional[str] = None

class LoanApplication(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    application_number: str
    customer_id: str
    loan_amount: float
    purpose: str
    guarantor_name: str
    guarantor_phone: str
    status: LoanStatus = LoanStatus.PENDING
    applied_by: str
    applied_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    reviewed_by: Optional[str] = None
    reviewed_at: Optional[datetime] = None
    notes: Optional[str] = None

class LoanApplicationCreate(BaseModel):
    customer_id: str
    loan_amount: float
    purpose: str
    guarantor_name: str
    guarantor_phone: str

class Loan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    loan_number: str
    application_id: str
    customer_id: str
    principal_amount: float
    interest_rate: float
    duration_months: int
    monthly_payment: float
    total_repayable: float
    amount_paid: float = 0.0
    outstanding_balance: float
    disbursement_date: Optional[datetime] = None
    status: LoanStatus = LoanStatus.APPROVED
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LoanCreate(BaseModel):
    application_id: str
    customer_id: str
    principal_amount: float
    interest_rate: float
    duration_months: int

class LoanRepayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    loan_id: str
    customer_id: str
    amount: float
    payment_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    received_by: str
    notes: Optional[str] = None

class LoanRepaymentCreate(BaseModel):
    loan_id: str
    customer_id: str
    amount: float
    notes: Optional[str] = None

class Investment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    investment_number: str
    investor_id: str
    amount: float
    plan: str
    interest_rate: float
    start_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    maturity_date: datetime
    is_active: bool = True

class InvestmentCreate(BaseModel):
    investor_id: str
    amount: float
    plan: str
    interest_rate: float
    duration_months: int

class ATMTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_ref: str
    card_type: CardType
    bank_name: str
    card_last_four: str
    transaction_type: ATMTransactionType
    amount: float
    service_fee: float
    operator_id: str
    transaction_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = None

class ATMTransactionCreate(BaseModel):
    card_type: CardType
    bank_name: str
    card_last_four: str
    transaction_type: ATMTransactionType
    amount: float
    service_fee: float
    notes: Optional[str] = None

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_code: str
    name: str
    barcode: Optional[str] = None
    cost_price: float
    selling_price: float
    quantity_in_stock: int
    reorder_level: int = 10
    category: Optional[str] = None

class ProductCreate(BaseModel):
    name: str
    barcode: Optional[str] = None
    cost_price: float
    selling_price: float
    quantity_in_stock: int
    reorder_level: int = 10
    category: Optional[str] = None

class Sale(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    receipt_number: str
    operator_id: str
    items: List[Dict[str, Any]]
    subtotal: float
    tax: float = 0.0
    total: float
    payment_method: str
    sale_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SaleCreate(BaseModel):
    items: List[Dict[str, Any]]
    payment_method: str

class CashLedgerEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    transaction_type: TransactionType
    amount: float
    description: str
    reference_id: Optional[str] = None
    recorded_by: str
    branch_id: str
    transaction_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CashLedgerCreate(BaseModel):
    transaction_type: TransactionType
    amount: float
    description: str
    reference_id: Optional[str] = None

class Branch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    address: str
    phone: str
    manager_id: Optional[str] = None
    is_active: bool = True

class BranchCreate(BaseModel):
    name: str
    address: str
    phone: str
    manager_id: Optional[str] = None

class DashboardMetrics(BaseModel):
    total_customers: int
    total_savings: float
    total_loans_disbursed: float
    total_loans_outstanding: float
    cash_on_hand: float
    daily_collections: float
    pending_loan_applications: int
    overdue_loans: int

# Helper Functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(hours=24)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        user_data = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user_data is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        if isinstance(user_data.get('created_at'), str):
            user_data['created_at'] = datetime.fromisoformat(user_data['created_at'])
        
        return User(**user_data)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

def serialize_datetime(obj):
    """Convert datetime objects to ISO format strings for MongoDB storage"""
    if isinstance(obj, dict):
        return {k: serialize_datetime(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [serialize_datetime(item) for item in obj]
    elif isinstance(obj, datetime):
        return obj.isoformat()
    return obj

def deserialize_datetime(obj, datetime_fields):
    """Convert ISO format strings back to datetime objects"""
    if isinstance(obj, dict):
        for field in datetime_fields:
            if field in obj and isinstance(obj[field], str):
                obj[field] = datetime.fromisoformat(obj[field])
    return obj

# Authentication Endpoints
@api_router.post("/auth/register", response_model=User)
async def register(user_input: UserCreate):
    existing_user = await db.users.find_one({"email": user_input.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = hash_password(user_input.password)
    user_dict = user_input.model_dump(exclude={"password"})
    user = User(**user_dict)
    
    doc = serialize_datetime(user.model_dump())
    doc['password'] = hashed_password
    
    await db.users.insert_one(doc)
    return user

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user_data = await db.users.find_one({"email": credentials.email})
    if not user_data:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(credentials.password, user_data.get('password', '')):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user_data.get('is_active', True):
        raise HTTPException(status_code=401, detail="Account is deactivated")
    
    user_data = deserialize_datetime(user_data, ['created_at'])
    user = User(**{k: v for k, v in user_data.items() if k != 'password' and k != '_id'})
    
    access_token = create_access_token(data={"sub": user.id})
    
    return TokenResponse(access_token=access_token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Customer Endpoints
@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_input: CustomerCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.BRANCH_MANAGER, UserRole.SAVINGS_OFFICER, UserRole.AGENT]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    count = await db.customers.count_documents({})
    customer_number = f"CUST{count + 1:06d}"
    
    customer_dict = customer_input.model_dump()
    customer = Customer(
        **customer_dict,
        customer_number=customer_number,
        agent_id=current_user.id,
        branch_id=current_user.branch_id or "default"
    )
    
    doc = serialize_datetime(customer.model_dump())
    await db.customers.insert_one(doc)
    return customer

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(
    agent_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if current_user.role == UserRole.AGENT:
        query['agent_id'] = current_user.id
    elif agent_id:
        query['agent_id'] = agent_id
    
    customers = await db.customers.find(query, {"_id": 0}).to_list(1000)
    for customer in customers:
        deserialize_datetime(customer, ['created_at'])
    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    deserialize_datetime(customer, ['created_at'])
    return Customer(**customer)

# Savings Endpoints
@api_router.post("/savings/accounts", response_model=SavingsAccount)
async def create_savings_account(account_input: SavingsAccountCreate, current_user: User = Depends(get_current_user)):
    count = await db.savings_accounts.count_documents({})
    account_number = f"SAV{count + 1:08d}"
    
    account = SavingsAccount(**account_input.model_dump(), account_number=account_number)
    doc = serialize_datetime(account.model_dump())
    await db.savings_accounts.insert_one(doc)
    return account

@api_router.get("/savings/accounts", response_model=List[SavingsAccount])
async def get_savings_accounts(customer_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if customer_id:
        query['customer_id'] = customer_id
    
    accounts = await db.savings_accounts.find(query, {"_id": 0}).to_list(1000)
    for account in accounts:
        deserialize_datetime(account, ['created_at'])
    return accounts

@api_router.post("/savings/transactions", response_model=SavingsTransaction)
async def create_savings_transaction(
    transaction_input: SavingsTransactionCreate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in [UserRole.AGENT, UserRole.SAVINGS_OFFICER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    account = await db.savings_accounts.find_one({"id": transaction_input.account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Savings account not found")
    
    transaction = SavingsTransaction(**transaction_input.model_dump(), agent_id=current_user.id)
    doc = serialize_datetime(transaction.model_dump())
    await db.savings_transactions.insert_one(doc)
    
    new_balance = account.get('current_balance', 0) + transaction_input.amount
    await db.savings_accounts.update_one(
        {"id": transaction_input.account_id},
        {"$set": {"current_balance": new_balance}}
    )
    
    return transaction

@api_router.get("/savings/transactions", response_model=List[SavingsTransaction])
async def get_savings_transactions(
    agent_id: Optional[str] = None,
    customer_id: Optional[str] = None,
    verified: Optional[bool] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if current_user.role == UserRole.AGENT:
        query['agent_id'] = current_user.id
    elif agent_id:
        query['agent_id'] = agent_id
    
    if customer_id:
        query['customer_id'] = customer_id
    if verified is not None:
        query['verified_by_cashier'] = verified
    
    transactions = await db.savings_transactions.find(query, {"_id": 0}).to_list(1000)
    for txn in transactions:
        deserialize_datetime(txn, ['collection_date', 'verification_date'])
    return transactions

@api_router.post("/savings/transactions/{transaction_id}/verify")
async def verify_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.CASHIER, UserRole.BRANCH_MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.savings_transactions.update_one(
        {"id": transaction_id},
        {"$set": {
            "verified_by_cashier": True,
            "cashier_id": current_user.id,
            "verification_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    return {"message": "Transaction verified successfully"}

# Loan Endpoints
@api_router.post("/loans/applications", response_model=LoanApplication)
async def create_loan_application(app_input: LoanApplicationCreate, current_user: User = Depends(get_current_user)):
    count = await db.loan_applications.count_documents({})
    application_number = f"LOAN{count + 1:06d}"
    
    application = LoanApplication(
        **app_input.model_dump(),
        application_number=application_number,
        applied_by=current_user.id
    )
    doc = serialize_datetime(application.model_dump())
    await db.loan_applications.insert_one(doc)
    return application

@api_router.get("/loans/applications", response_model=List[LoanApplication])
async def get_loan_applications(
    status: Optional[LoanStatus] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if status:
        query['status'] = status
    
    applications = await db.loan_applications.find(query, {"_id": 0}).to_list(1000)
    for app in applications:
        deserialize_datetime(app, ['applied_at', 'reviewed_at'])
    return applications

@api_router.post("/loans/applications/{application_id}/approve")
async def approve_loan_application(application_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.LOAN_OFFICER, UserRole.BRANCH_MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.loan_applications.update_one(
        {"id": application_id},
        {"$set": {
            "status": LoanStatus.APPROVED.value,
            "reviewed_by": current_user.id,
            "reviewed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Application not found")
    
    return {"message": "Loan application approved"}

@api_router.post("/loans", response_model=Loan)
async def create_loan(loan_input: LoanCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.LOAN_OFFICER, UserRole.BRANCH_MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    count = await db.loans.count_documents({})
    loan_number = f"LN{count + 1:08d}"
    
    monthly_rate = loan_input.interest_rate / 100 / 12
    monthly_payment = (loan_input.principal_amount * monthly_rate * (1 + monthly_rate) ** loan_input.duration_months) / ((1 + monthly_rate) ** loan_input.duration_months - 1)
    total_repayable = monthly_payment * loan_input.duration_months
    
    loan = Loan(
        **loan_input.model_dump(),
        loan_number=loan_number,
        monthly_payment=round(monthly_payment, 2),
        total_repayable=round(total_repayable, 2),
        outstanding_balance=round(total_repayable, 2)
    )
    doc = serialize_datetime(loan.model_dump())
    await db.loans.insert_one(doc)
    
    await db.loan_applications.update_one(
        {"id": loan_input.application_id},
        {"$set": {"status": LoanStatus.DISBURSED.value}}
    )
    
    return loan

@api_router.get("/loans", response_model=List[Loan])
async def get_loans(customer_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if customer_id:
        query['customer_id'] = customer_id
    
    loans = await db.loans.find(query, {"_id": 0}).to_list(1000)
    for loan in loans:
        deserialize_datetime(loan, ['disbursement_date', 'created_at'])
    return loans

@api_router.post("/loans/repayments", response_model=LoanRepayment)
async def create_loan_repayment(repayment_input: LoanRepaymentCreate, current_user: User = Depends(get_current_user)):
    loan = await db.loans.find_one({"id": repayment_input.loan_id})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    repayment = LoanRepayment(**repayment_input.model_dump(), received_by=current_user.id)
    doc = serialize_datetime(repayment.model_dump())
    await db.loan_repayments.insert_one(doc)
    
    new_amount_paid = loan.get('amount_paid', 0) + repayment_input.amount
    new_outstanding = loan.get('outstanding_balance', 0) - repayment_input.amount
    
    update_data = {
        "amount_paid": new_amount_paid,
        "outstanding_balance": max(0, new_outstanding)
    }
    if new_outstanding <= 0:
        update_data['status'] = LoanStatus.COMPLETED.value
    
    await db.loans.update_one({"id": repayment_input.loan_id}, {"$set": update_data})
    
    return repayment

@api_router.get("/loans/repayments", response_model=List[LoanRepayment])
async def get_loan_repayments(loan_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if loan_id:
        query['loan_id'] = loan_id
    
    repayments = await db.loan_repayments.find(query, {"_id": 0}).to_list(1000)
    for rep in repayments:
        deserialize_datetime(rep, ['payment_date'])
    return repayments

# Investment Endpoints
@api_router.post("/investments", response_model=Investment)
async def create_investment(investment_input: InvestmentCreate, current_user: User = Depends(get_current_user)):
    count = await db.investments.count_documents({})
    investment_number = f"INV{count + 1:06d}"
    
    maturity_date = datetime.now(timezone.utc) + timedelta(days=investment_input.duration_months * 30)
    
    investment = Investment(
        **investment_input.model_dump(exclude={'duration_months'}),
        investment_number=investment_number,
        maturity_date=maturity_date
    )
    doc = serialize_datetime(investment.model_dump())
    await db.investments.insert_one(doc)
    return investment

@api_router.get("/investments", response_model=List[Investment])
async def get_investments(investor_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.role == UserRole.INVESTOR:
        query['investor_id'] = current_user.id
    elif investor_id:
        query['investor_id'] = investor_id
    
    investments = await db.investments.find(query, {"_id": 0}).to_list(1000)
    for inv in investments:
        deserialize_datetime(inv, ['start_date', 'maturity_date'])
    return investments

# ATM Transaction Endpoints
@api_router.post("/atm/transactions", response_model=ATMTransaction)
async def create_atm_transaction(txn_input: ATMTransactionCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.POS_OPERATOR, UserRole.CASHIER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    count = await db.atm_transactions.count_documents({})
    transaction_ref = f"ATM{count + 1:08d}"
    
    transaction = ATMTransaction(**txn_input.model_dump(), transaction_ref=transaction_ref, operator_id=current_user.id)
    doc = serialize_datetime(transaction.model_dump())
    await db.atm_transactions.insert_one(doc)
    
    txn_type = TransactionType.DEPOSIT if txn_input.transaction_type == ATMTransactionType.DEPOSIT else TransactionType.WITHDRAWAL
    ledger_entry = CashLedgerEntry(
        transaction_type=txn_type,
        amount=txn_input.amount if txn_input.transaction_type == ATMTransactionType.DEPOSIT else -txn_input.amount,
        description=f"ATM {txn_input.transaction_type.value} - {transaction_ref}",
        reference_id=transaction.id,
        recorded_by=current_user.id,
        branch_id=current_user.branch_id or "default"
    )
    await db.cash_ledger.insert_one(serialize_datetime(ledger_entry.model_dump()))
    
    return transaction

@api_router.get("/atm/transactions", response_model=List[ATMTransaction])
async def get_atm_transactions(current_user: User = Depends(get_current_user)):
    transactions = await db.atm_transactions.find({}, {"_id": 0}).to_list(1000)
    for txn in transactions:
        deserialize_datetime(txn, ['transaction_date'])
    return transactions

# Product Endpoints
@api_router.post("/products", response_model=Product)
async def create_product(product_input: ProductCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.BRANCH_MANAGER, UserRole.POS_OPERATOR]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    count = await db.products.count_documents({})
    product_code = f"PROD{count + 1:06d}"
    
    product = Product(**product_input.model_dump(), product_code=product_code)
    await db.products.insert_one(product.model_dump())
    return product

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    return products

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, quantity_change: int, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    new_quantity = product.get('quantity_in_stock', 0) + quantity_change
    await db.products.update_one({"id": product_id}, {"$set": {"quantity_in_stock": new_quantity}})
    return {"message": "Product updated", "new_quantity": new_quantity}

# Sales Endpoints
@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_input: SaleCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in [UserRole.POS_OPERATOR, UserRole.CASHIER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    count = await db.sales.count_documents({})
    receipt_number = f"RCPT{count + 1:08d}"
    
    subtotal = sum(item['quantity'] * item['price'] for item in sale_input.items)
    total = subtotal
    
    sale = Sale(
        **sale_input.model_dump(),
        receipt_number=receipt_number,
        operator_id=current_user.id,
        subtotal=subtotal,
        total=total
    )
    doc = serialize_datetime(sale.model_dump())
    await db.sales.insert_one(doc)
    
    for item in sale_input.items:
        await db.products.update_one(
            {"id": item['product_id']},
            {"$inc": {"quantity_in_stock": -item['quantity']}}
        )
    
    ledger_entry = CashLedgerEntry(
        transaction_type=TransactionType.SALE,
        amount=total,
        description=f"POS Sale - {receipt_number}",
        reference_id=sale.id,
        recorded_by=current_user.id,
        branch_id=current_user.branch_id or "default"
    )
    await db.cash_ledger.insert_one(serialize_datetime(ledger_entry.model_dump()))
    
    return sale

@api_router.get("/sales", response_model=List[Sale])
async def get_sales(current_user: User = Depends(get_current_user)):
    sales = await db.sales.find({}, {"_id": 0}).to_list(1000)
    for sale in sales:
        deserialize_datetime(sale, ['sale_date'])
    return sales

# Cash Ledger Endpoints
@api_router.get("/cash/ledger", response_model=List[CashLedgerEntry])
async def get_cash_ledger(current_user: User = Depends(get_current_user)):
    entries = await db.cash_ledger.find({}, {"_id": 0}).to_list(1000)
    for entry in entries:
        deserialize_datetime(entry, ['transaction_date'])
    return entries

@api_router.get("/cash/balance")
async def get_cash_balance(current_user: User = Depends(get_current_user)):
    entries = await db.cash_ledger.find({}).to_list(10000)
    balance = sum(entry.get('amount', 0) for entry in entries)
    return {"balance": balance}

# Dashboard Endpoints
@api_router.get("/dashboard/metrics", response_model=DashboardMetrics)
async def get_dashboard_metrics(current_user: User = Depends(get_current_user)):
    total_customers = await db.customers.count_documents({})
    
    savings_accounts = await db.savings_accounts.find({}).to_list(10000)
    total_savings = sum(acc.get('current_balance', 0) for acc in savings_accounts)
    
    loans = await db.loans.find({}).to_list(10000)
    total_loans_disbursed = sum(loan.get('principal_amount', 0) for loan in loans)
    total_loans_outstanding = sum(loan.get('outstanding_balance', 0) for loan in loans)
    
    ledger_entries = await db.cash_ledger.find({}).to_list(10000)
    cash_on_hand = sum(entry.get('amount', 0) for entry in ledger_entries)
    
    today = datetime.now(timezone.utc).date()
    daily_collections = 0
    for entry in ledger_entries:
        entry_date = entry.get('transaction_date')
        if isinstance(entry_date, str):
            entry_date = datetime.fromisoformat(entry_date)
        if entry_date and entry_date.date() == today:
            daily_collections += entry.get('amount', 0)
    
    pending_loan_applications = await db.loan_applications.count_documents({"status": LoanStatus.PENDING.value})
    
    overdue_loans = 0
    
    return DashboardMetrics(
        total_customers=total_customers,
        total_savings=total_savings,
        total_loans_disbursed=total_loans_disbursed,
        total_loans_outstanding=total_loans_outstanding,
        cash_on_hand=cash_on_hand,
        daily_collections=daily_collections,
        pending_loan_applications=pending_loan_applications,
        overdue_loans=overdue_loans
    )

# Branch Endpoints
@api_router.post("/branches", response_model=Branch)
async def create_branch(branch_input: BranchCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    branch = Branch(**branch_input.model_dump())
    await db.branches.insert_one(branch.model_dump())
    return branch

@api_router.get("/branches", response_model=List[Branch])
async def get_branches(current_user: User = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    return branches

# Report Generation Endpoints
@api_router.get("/reports/agent-collections/excel")
async def export_agent_collections_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if start_date and end_date:
        query['collection_date'] = {
            "$gte": start_date,
            "$lte": end_date
        }
    
    transactions = await db.savings_transactions.find(query, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agent Collections"
    
    headers = ["Transaction ID", "Agent ID", "Customer ID", "Amount", "Collection Date", "Verified"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for txn in transactions:
        collection_date = txn.get('collection_date')
        if isinstance(collection_date, str):
            collection_date = datetime.fromisoformat(collection_date)
        ws.append([
            txn.get('id'),
            txn.get('agent_id'),
            txn.get('customer_id'),
            txn.get('amount'),
            collection_date.strftime("%Y-%m-%d %H:%M") if collection_date else "",
            "Yes" if txn.get('verified_by_cashier') else "No"
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=agent_collections.xlsx"}
    )

@api_router.get("/reports/loan-portfolio/excel")
async def export_loan_portfolio_excel(current_user: User = Depends(get_current_user)):
    loans = await db.loans.find({}, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Portfolio"
    
    headers = ["Loan Number", "Customer ID", "Principal", "Interest Rate", "Total Repayable", "Amount Paid", "Outstanding", "Status"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for loan in loans:
        ws.append([
            loan.get('loan_number'),
            loan.get('customer_id'),
            loan.get('principal_amount'),
            f"{loan.get('interest_rate')}%",
            loan.get('total_repayable'),
            loan.get('amount_paid'),
            loan.get('outstanding_balance'),
            loan.get('status')
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=loan_portfolio.xlsx"}
    )

@api_router.get("/")
async def root():
    return {"message": "NaijaFinance API", "version": "1.0"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
